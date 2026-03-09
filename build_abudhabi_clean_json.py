import json
from openpyxl import load_workbook

INPUT_FILE = "knowledge/Resturants Bars- AbuDhabi.xlsx"
OUTPUT_FILE = "knowledge/abudhabi_restaurants_clean.json"


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_float(value):
    if value is None:
        return None
    try:
        return float(value)
    except:
        return None


def build_clean_json():
    workbook = load_workbook(INPUT_FILE, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    records = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        tourism_license = clean_text(row[0])
        establishment_name = clean_text(row[1])
        location = clean_text(row[2])
        latitude = clean_float(row[3])
        longitude = clean_float(row[4])

        # skip rows that are completely empty
        if not tourism_license and not establishment_name and not location:
            continue

        record = {
            "tourism_license": tourism_license,
            "establishment_name": establishment_name,
            "location": location,
            "latitude": latitude,
            "longitude": longitude
        }

        records.append(record)

    final_data = {
        "city": "Abu Dhabi",
        "source": "Abu Dhabi Restaurant Tourism Licenses Dataset 2025",
        "record_count": len(records),
        "restaurants": records
    }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(final_data, f, ensure_ascii=False, indent=2)

    print(f"Done. Clean JSON saved to {OUTPUT_FILE}")
    print(f"Total records: {len(records)}")


if __name__ == "__main__":
    build_clean_json()