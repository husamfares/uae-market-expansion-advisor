from openpyxl import load_workbook


FILE_NAME = "knowledge/Resturants Bars- AbuDhabi.xlsx"

def inspect_excel():
    workbook = load_workbook(FILE_NAME, data_only=True)

    print("Sheet names:")
    print(workbook.sheetnames)

    sheet = workbook[workbook.sheetnames[0]]

    print("\nUsing sheet:", sheet.title)

    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)

    print("\nColumns:")
    for i, header in enumerate(headers, start=1):
        print(f"{i}. {header}")

    print("\nFirst 5 data rows:")
    for row in sheet.iter_rows(min_row=2, max_row=6, values_only=True):
        print(row)


if __name__ == "__main__":
    inspect_excel()